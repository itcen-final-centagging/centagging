"""프롬프트 평가 CSV를 여러 시트의 Excel 보고서로 변환합니다."""

import argparse
import csv
import math
import pathlib
import statistics
import sys
from collections import defaultdict
from collections.abc import Mapping, Sequence

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

CsvRow = dict[str, object]

PROMPT_LABELS = {
    "detection": "객체 탐지",
    "attribute": "속성 추출",
    "xai": "XAI",
}
GROUP_LABELS = {"common": "공통", "quality": "품질", "efficiency": "효율"}
METRIC_LABELS = {
    "success_rate": "정상 응답률",
    "average_ms": "평균 처리 시간",
    "p95_ms": "P95 처리 시간",
    "batch_p95_ms": "배치 P95 처리 시간",
    "total_tokens": "전체 토큰",
    "tokens_per_success": "성공 1건당 토큰",
    "retry_rate": "재시도율",
    "token_usage_coverage": "토큰 계측률",
    "f1": "F1",
    "mean_iou": "평균 IoU",
    "miss_rate": "미탐률",
    "korean_evidence_rate": "한글 근거 비율",
    "confidence_present_rate": "신뢰도 응답률",
    "correct_objects_per_1k_tokens": "1,000토큰당 정탐 객체",
    "ms_per_correct_object": "정탐 객체당 처리 시간",
    "missing_rate": "속성 누락률",
    "incorrect_pair_rate": "잘못된 속성 비율",
    "category_accuracy": "카테고리 정확도",
    "sub_category_accuracy": "서브 카테고리 정확도",
    "correct_pairs_per_1k_tokens": "1,000토큰당 정답 속성",
    "ms_per_correct_pair": "정답 속성당 처리 시간",
    "crop_coverage_rate": "Crop 반환률",
    "sku_exact_once_rate": "SKU 단일 반환률",
    "criteria_complete_rate": "루브릭 완전성",
    "criteria_range_valid_rate": "루브릭 점수 범위 준수율",
    "score_sum_valid_rate": "총점 합계 일치율",
    "status_threshold_valid_rate": "판정 임계값 준수율",
    "object_label_present_rate": "객체 라벨 응답률",
    "object_label_match_rate": "객체 라벨 일치율",
    "mood_present_rate": "VLM mood 응답률",
}
LOWER_IS_BETTER = {
    "average_ms",
    "p95_ms",
    "batch_p95_ms",
    "total_tokens",
    "tokens_per_success",
    "retry_rate",
    "miss_rate",
    "missing_rate",
    "incorrect_pair_rate",
    "ms_per_correct_object",
    "ms_per_correct_pair",
}
RATIO_METRICS = {
    "success_rate",
    "retry_rate",
    "token_usage_coverage",
    "f1",
    "mean_iou",
    "miss_rate",
    "korean_evidence_rate",
    "confidence_present_rate",
    "missing_rate",
    "incorrect_pair_rate",
    "category_accuracy",
    "sub_category_accuracy",
    "crop_coverage_rate",
    "sku_exact_once_rate",
    "criteria_complete_rate",
    "criteria_range_valid_rate",
    "score_sum_valid_rate",
    "status_threshold_valid_rate",
    "object_label_present_rate",
    "object_label_match_rate",
    "mood_present_rate",
}
XAI_QUALITY_METRICS = (
    "crop_coverage_rate",
    "sku_exact_once_rate",
    "criteria_complete_rate",
    "criteria_range_valid_rate",
    "score_sum_valid_rate",
    "status_threshold_valid_rate",
    "object_label_present_rate",
    "object_label_match_rate",
    "mood_present_rate",
)

NAVY = "17365D"
BLUE = "2F75B5"
LIGHT_BLUE = "D9EAF7"
PALE_BLUE = "EAF2F8"
GREEN = "E2F0D9"
RED = "FCE4D6"
GRAY = "E7E6E6"
WHITE = "FFFFFF"
THIN_GRAY = Side(style="thin", color="D9E2F3")


def read_csv(path: pathlib.Path) -> list[dict[str, str]]:
    """UTF-8 BOM CSV를 읽고 빈 행을 제외합니다."""
    with path.open(encoding="utf-8-sig", newline="") as source:
        return [row for row in csv.DictReader(source) if any(row.values())]


def to_number(value: object) -> float | None:
    """CSV 값을 실수로 변환하며 빈 값은 유지합니다."""
    if value in (None, ""):
        return None
    try:
        return float(str(value))
    except ValueError:
        return None


def calculate_p95(values: Sequence[float]) -> float:
    """표본 수가 적어도 재현 가능한 최근접 순위 P95를 계산합니다."""
    if not values:
        return 0.0
    ordered = sorted(values)
    return round(ordered[max(math.ceil(len(ordered) * 0.95) - 1, 0)], 6)


# pylint: disable-next=too-many-locals
def build_xai_summary_rows(rows: Sequence[Mapping[str, str]]) -> list[CsvRow]:
    """XAI 호출별 결과를 v1·v2 공통/품질 지표로 집계합니다."""
    grouped: dict[str, list[Mapping[str, str]]] = defaultdict(list)
    for row in rows:
        grouped[str(row.get("prompt_version", ""))].append(row)

    summary_rows: list[CsvRow] = []
    for version in sorted(grouped):
        version_rows = grouped[version]
        successful = [
            row for row in version_rows if to_number(row.get("success")) == 1
        ]
        durations = [
            number
            for row in successful
            if (number := to_number(row.get("duration_ms"))) is not None
        ]
        definitions: list[tuple[str, str, float, str]] = [
            (
                "common",
                "success_rate",
                round(len(successful) / len(version_rows), 6),
                "ratio",
            ),
            (
                "common",
                "average_ms",
                round(statistics.fmean(durations), 6) if durations else 0.0,
                "ms",
            ),
            ("common", "p95_ms", calculate_p95(durations), "ms"),
        ]
        for metric in XAI_QUALITY_METRICS:
            values = [
                number
                for row in successful
                if (number := to_number(row.get(metric))) is not None
            ]
            definitions.append(
                (
                    "quality",
                    metric,
                    round(statistics.fmean(values), 6) if values else 0.0,
                    "ratio",
                )
            )

        first = version_rows[0]
        for metric_group, metric, value, unit in definitions:
            summary_rows.append(
                {
                    "evaluated_at": first.get("evaluated_at", ""),
                    "dataset_name": "xai-prompt-evaluation-v1-v2",
                    "model": first.get("model", ""),
                    "provider": "",
                    "location": "",
                    "prompt_type": "xai",
                    "prompt_version": version,
                    "metric_group": metric_group,
                    "metric": metric,
                    "value": value,
                    "unit": unit,
                    "target": "",
                    "direction": (
                        "lower" if metric in LOWER_IS_BETTER else "higher"
                    ),
                    "passed": "",
                }
            )
    return summary_rows


def build_comparison_rows(
    summary_rows: Sequence[Mapping[str, object]],
) -> list[CsvRow]:
    """동일 지표의 v1·v2 값과 우세 버전을 계산합니다."""
    index = {
        (
            str(row.get("prompt_type", "")),
            str(row.get("metric_group", "")),
            str(row.get("metric", "")),
            str(row.get("prompt_version", "")),
        ): row
        for row in summary_rows
    }
    keys = {key[:3] for key in index if key[3] in {"v1", "v2"}}
    comparison: list[CsvRow] = []
    for prompt_type, metric_group, metric in sorted(keys):
        v1_row = index.get((prompt_type, metric_group, metric, "v1"))
        v2_row = index.get((prompt_type, metric_group, metric, "v2"))
        if not v1_row or not v2_row:
            continue
        v1_value = to_number(v1_row.get("value"))
        v2_value = to_number(v2_row.get("value"))
        if v1_value is None or v2_value is None:
            continue
        direction = str(v2_row.get("direction") or "higher")
        delta = v2_value - v1_value
        if math.isclose(v1_value, v2_value):
            winner = "tie"
        elif direction == "lower":
            winner = "v2" if v2_value < v1_value else "v1"
        else:
            winner = "v2" if v2_value > v1_value else "v1"
        comparison.append(
            {
                "dataset_name": v2_row.get("dataset_name", ""),
                "model": v2_row.get("model", ""),
                "prompt_type": prompt_type,
                "metric_group": metric_group,
                "metric": metric,
                "v1_value": round(v1_value, 6),
                "v2_value": round(v2_value, 6),
                "unit": v2_row.get("unit", ""),
                "target": v2_row.get("target", ""),
                "direction": direction,
                "delta": round(delta, 6),
                "change_percent": (
                    round(delta / v1_value * 100, 2) if v1_value else ""
                ),
                "winner": winner,
            }
        )
    return comparison


def translated(value: object, translations: Mapping[str, str]) -> str:
    """번역 사전에 없는 값은 원문을 유지합니다."""
    key = str(value)
    return translations.get(key, key)


def add_title(sheet, title: str, subtitle: str) -> None:
    """시트 상단에 공통 제목 영역을 만듭니다."""
    sheet.merge_cells("A1:L1")
    sheet["A1"] = title
    sheet["A1"].font = Font(name="맑은 고딕", size=18, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet["A1"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 34
    sheet.merge_cells("A2:L2")
    sheet["A2"] = subtitle
    sheet["A2"].font = Font(name="맑은 고딕", size=10, color="44546A")
    sheet["A2"].fill = PatternFill("solid", fgColor=PALE_BLUE)
    sheet["A2"].alignment = Alignment(vertical="center")
    sheet.row_dimensions[2].height = 24
    sheet.sheet_view.showGridLines = False


def write_table(
    sheet,
    start_row: int,
    headers: Sequence[str],
    rows: Sequence[Sequence[object]],
    table_name: str,
) -> int:
    """값과 기본 스타일을 포함한 Excel 테이블을 작성합니다."""
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(start_row, column, header)
        cell.font = Font(name="맑은 고딕", bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=THIN_GRAY)
    for row_index, values in enumerate(rows, start=start_row + 1):
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_index, column, value)
            cell.font = Font(name="맑은 고딕", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=THIN_GRAY)
    last_row = start_row + len(rows)
    if rows:
        last_column = get_column_letter(len(headers))
        table = Table(
            displayName=table_name,
            ref=f"A{start_row}:{last_column}{last_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    return last_row


def comparison_values(
    rows: Sequence[Mapping[str, object]],
) -> list[list[object]]:
    """비교 행을 사람이 읽기 쉬운 열 순서로 변환합니다."""
    return [
        [
            translated(row.get("prompt_type"), PROMPT_LABELS),
            translated(row.get("metric_group"), GROUP_LABELS),
            translated(row.get("metric"), METRIC_LABELS),
            to_number(row.get("v1_value")),
            to_number(row.get("v2_value")),
            row.get("unit", ""),
            to_number(row.get("delta")),
            to_number(row.get("change_percent")),
            row.get("direction", ""),
            row.get("winner", ""),
            to_number(row.get("target")),
        ]
        for row in rows
    ]


def format_comparison(
    sheet, rows: Sequence[Mapping[str, object]], start_row: int
) -> None:
    """비교 테이블에 단위별 숫자 형식과 승자 색상을 적용합니다."""
    for excel_row, row in enumerate(rows, start=start_row + 1):
        metric = str(row.get("metric", ""))
        number_format = "0.00%" if metric in RATIO_METRICS else "#,##0.00"
        for column in (4, 5, 7, 11):
            sheet.cell(excel_row, column).number_format = number_format
        sheet.cell(excel_row, 8).number_format = "0.00"
    if not rows:
        return
    winner_range = f"J{start_row + 1}:J{start_row + len(rows)}"
    for winner, color in (("v2", GREEN), ("v1", RED), ("tie", GRAY)):
        sheet.conditional_formatting.add(
            winner_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{winner}"'],
                fill=PatternFill("solid", fgColor=color),
            ),
        )


def set_standard_widths(sheet) -> None:
    """비교 시트의 열 너비를 읽기 좋은 크기로 제한합니다."""
    widths = (15, 11, 28, 14, 14, 18, 14, 14, 12, 12, 14, 3)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def build_overview_sheet(workbook: Workbook, rows: Sequence[CsvRow]) -> None:
    """세 프롬프트의 전체 비교와 승자 수를 정리합니다."""
    sheet = workbook.create_sheet("종합 비교")
    add_title(
        sheet,
        "프롬프트 v1 / v2 품질 비교",
        "객체 탐지·속성 추출·XAI의 공통 지표와 고유 품질 지표입니다.",
    )
    for cell, label in (
        ("A4", "v2 우세"),
        ("D4", "v1 우세"),
        ("G4", "동률"),
        ("J4", "전체 지표"),
    ):
        sheet[cell] = label
        sheet[cell].font = Font(name="맑은 고딕", bold=True, color="44546A")
        sheet[cell].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        sheet[cell].alignment = Alignment(horizontal="center")
    table_start = 8
    table_end = table_start + len(rows)
    sheet["A5"] = f'=COUNTIF(J{table_start + 1}:J{table_end},"v2")'
    sheet["D5"] = f'=COUNTIF(J{table_start + 1}:J{table_end},"v1")'
    sheet["G5"] = f'=COUNTIF(J{table_start + 1}:J{table_end},"tie")'
    sheet["J5"] = f"=COUNTA(C{table_start + 1}:C{table_end})"
    for cell in ("A5", "D5", "G5", "J5"):
        sheet[cell].font = Font(
            name="맑은 고딕", size=18, bold=True, color=NAVY
        )
        sheet[cell].alignment = Alignment(horizontal="center")
    headers = (
        "프롬프트",
        "지표 구분",
        "지표",
        "v1",
        "v2",
        "단위",
        "증감",
        "변화율(%)",
        "판정 방향",
        "우세 버전",
        "목표",
    )
    write_table(
        sheet,
        table_start,
        headers,
        comparison_values(rows),
        "OverviewComparisonTable",
    )
    format_comparison(sheet, rows, table_start)
    sheet.freeze_panes = f"A{table_start + 1}"
    set_standard_widths(sheet)


# pylint: disable-next=too-many-arguments,too-many-positional-arguments,too-many-locals
def build_metric_sheet(
    workbook: Workbook,
    sheet_name: str,
    title: str,
    subtitle: str,
    comparison_rows: Sequence[CsvRow],
    summary_rows: Sequence[CsvRow],
    table_prefix: str,
) -> None:
    """프롬프트별 비교와 버전별 요약을 같은 시트에 작성합니다."""
    sheet = workbook.create_sheet(sheet_name)
    add_title(sheet, title, subtitle)
    headers = (
        "프롬프트",
        "지표 구분",
        "지표",
        "v1",
        "v2",
        "단위",
        "증감",
        "변화율(%)",
        "판정 방향",
        "우세 버전",
        "목표",
    )
    comparison_start = 4
    comparison_end = write_table(
        sheet,
        comparison_start,
        headers,
        comparison_values(comparison_rows),
        f"{table_prefix}ComparisonTable",
    )
    format_comparison(sheet, comparison_rows, comparison_start)

    summary_start = comparison_end + 3
    sheet.cell(summary_start - 1, 1, "버전별 지표 원본").font = Font(
        name="맑은 고딕", size=12, bold=True, color=NAVY
    )
    summary_headers = (
        "프롬프트",
        "버전",
        "지표 구분",
        "지표",
        "값",
        "단위",
        "목표",
        "판정 방향",
        "통과 여부",
    )
    summary_values = [
        [
            translated(row.get("prompt_type"), PROMPT_LABELS),
            row.get("prompt_version", ""),
            translated(row.get("metric_group"), GROUP_LABELS),
            translated(row.get("metric"), METRIC_LABELS),
            to_number(row.get("value")),
            row.get("unit", ""),
            to_number(row.get("target")),
            row.get("direction", ""),
            row.get("passed", ""),
        ]
        for row in summary_rows
    ]
    write_table(
        sheet,
        summary_start,
        summary_headers,
        summary_values,
        f"{table_prefix}SummaryTable",
    )
    for excel_row, row in enumerate(summary_rows, start=summary_start + 1):
        metric = str(row.get("metric", ""))
        number_format = "0.00%" if metric in RATIO_METRICS else "#,##0.00"
        sheet.cell(excel_row, 5).number_format = number_format
        sheet.cell(excel_row, 7).number_format = number_format
    sheet.freeze_panes = f"A{comparison_start + 1}"
    set_standard_widths(sheet)


def write_raw_section(
    sheet,
    start_row: int,
    title: str,
    rows: Sequence[Mapping[str, object]],
    table_name: str,
) -> int:
    """원본 CSV 한 종류를 원본 결과 시트에 기록합니다."""
    sheet.cell(start_row, 1, title).font = Font(
        name="맑은 고딕", size=12, bold=True, color=NAVY
    )
    if not rows:
        sheet.cell(start_row + 1, 1, "데이터 없음")
        return start_row + 3
    headers = list(rows[0].keys())
    values = [[row.get(header, "") for header in headers] for row in rows]
    return (
        write_table(
            sheet,
            start_row + 1,
            headers,
            values,
            table_name,
        )
        + 2
    )


def build_raw_sheet(
    workbook: Workbook,
    summary_rows: Sequence[Mapping[str, object]],
    comparison_rows: Sequence[Mapping[str, object]],
    xai_rows: Sequence[Mapping[str, object]],
) -> None:
    """입력 CSV 원본을 재검증할 수 있도록 한 시트에 보관합니다."""
    sheet = workbook.create_sheet("원본 결과")
    add_title(
        sheet,
        "평가 원본 결과",
        "요약 값의 근거가 되는 입력 CSV를 그대로 확인할 수 있습니다.",
    )
    next_row = write_raw_section(
        sheet, 4, "프롬프트 지표 요약", summary_rows, "RawSummaryTable"
    )
    next_row = write_raw_section(
        sheet,
        next_row,
        "프롬프트 버전 비교",
        comparison_rows,
        "RawComparisonTable",
    )
    write_raw_section(
        sheet, next_row, "XAI 호출별 결과", xai_rows, "RawXaiTable"
    )
    sheet.freeze_panes = "A5"
    for column_index in range(1, sheet.max_column + 1):
        max_length = max(
            len(str(sheet.cell(row, column_index).value or ""))
            for row in range(1, min(sheet.max_row, 200) + 1)
        )
        sheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_length + 2, 10), 32
        )


# pylint: disable-next=too-many-locals
def generate_workbook(
    metrics_summary_path: pathlib.Path,
    version_comparison_path: pathlib.Path,
    xai_results_path: pathlib.Path,
    output_path: pathlib.Path,
) -> pathlib.Path:
    """세 프롬프트의 평가 결과를 하나의 XLSX 보고서로 생성합니다."""
    base_summary: list[CsvRow] = [
        dict(row) for row in read_csv(metrics_summary_path)
    ]
    base_comparison: list[CsvRow] = [
        dict(row) for row in read_csv(version_comparison_path)
    ]
    xai_raw = read_csv(xai_results_path)
    xai_summary = build_xai_summary_rows(xai_raw)
    xai_comparison = build_comparison_rows(xai_summary)
    all_summary = [*base_summary, *xai_summary]
    all_comparison = [*base_comparison, *xai_comparison]

    workbook = Workbook()
    workbook.remove(workbook.active)
    build_overview_sheet(workbook, all_comparison)
    build_metric_sheet(
        workbook,
        "공통 지표",
        "공통 성능 지표",
        (
            "세 프롬프트를 정상 응답률과 처리 시간으로 비교합니다. "
            "토큰·재시도는 계측 가능한 프롬프트에만 표시됩니다."
        ),
        [row for row in all_comparison if row.get("metric_group") == "common"],
        [row for row in all_summary if row.get("metric_group") == "common"],
        "Common",
    )
    sheet_definitions = (
        (
            "detection",
            "객체 탐지",
            "객체 탐지 프롬프트",
            "탐지 정확도·박스 IoU·근거 언어·토큰 효율을 비교합니다.",
            "Detection",
        ),
        (
            "attribute",
            "속성 추출",
            "속성 추출 프롬프트",
            "속성 정확도·누락률·카테고리 정확도·토큰 효율을 비교합니다.",
            "Attribute",
        ),
        (
            "xai",
            "XAI",
            "XAI 프롬프트",
            "응답 계약·루브릭 점수·객체 라벨·VLM mood 품질을 비교합니다.",
            "Xai",
        ),
    )
    for prompt_type, sheet_name, title, subtitle, prefix in sheet_definitions:
        build_metric_sheet(
            workbook,
            sheet_name,
            title,
            subtitle,
            [
                row
                for row in all_comparison
                if row.get("prompt_type") == prompt_type
            ],
            [
                row
                for row in all_summary
                if row.get("prompt_type") == prompt_type
            ],
            prefix,
        )
    build_raw_sheet(workbook, base_summary, base_comparison, xai_raw)
    workbook.active = 0
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    """통합 보고서 입력·출력 경로를 파싱합니다."""
    parser = argparse.ArgumentParser()
    parser.add_argument("--metrics-summary", type=pathlib.Path, required=True)
    parser.add_argument(
        "--version-comparison", type=pathlib.Path, required=True
    )
    parser.add_argument("--xai-results", type=pathlib.Path, required=True)
    parser.add_argument("--output", type=pathlib.Path, required=True)
    return parser.parse_args()


def main() -> int:
    """명령행 인자로 전달된 CSV를 통합 XLSX로 변환합니다."""
    args = parse_args()
    for path in (
        args.metrics_summary,
        args.version_comparison,
        args.xai_results,
    ):
        if not path.is_file():
            print(f"입력 CSV가 없습니다: {path}", file=sys.stderr)
            return 2
    output_path = generate_workbook(
        args.metrics_summary,
        args.version_comparison,
        args.xai_results,
        args.output,
    )
    print(output_path.resolve())
    return 0


if __name__ == "__main__":
    sys.exit(main())
