"""프롬프트 평가 Excel 보고서 생성을 검증합니다."""

import csv
import pathlib
import tempfile
import unittest

from openpyxl import load_workbook

from app.evaluation import prompt_excel_report


def _write_csv(
    path: pathlib.Path,
    fieldnames: list[str],
    rows: list[dict[str, object]],
) -> None:
    """테스트 입력 CSV를 UTF-8 BOM으로 저장합니다."""
    with path.open("w", encoding="utf-8-sig", newline="") as target:
        writer = csv.DictWriter(target, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


class PromptExcelReportTest(unittest.TestCase):
    """세 프롬프트 결과가 하나의 통합 문서로 생성되는지 확인합니다."""

    def test_generate_workbook(self) -> None:
        """필수 시트와 XAI 집계 결과를 생성합니다."""
        with tempfile.TemporaryDirectory() as directory:
            root = pathlib.Path(directory)
            summary_path = root / "prompt_metrics_summary.csv"
            comparison_path = root / "prompt_version_comparison.csv"
            xai_path = root / "xai.csv"
            output_path = root / "prompt_evaluation_report.xlsx"

            summary_fields = [
                "evaluated_at",
                "dataset_name",
                "model",
                "provider",
                "location",
                "prompt_type",
                "prompt_version",
                "metric_group",
                "metric",
                "value",
                "unit",
                "target",
                "direction",
                "passed",
            ]
            summary_rows = []
            for prompt_type in ("detection", "attribute"):
                for version, value in (("v1", 0.8), ("v2", 0.9)):
                    summary_rows.append(
                        {
                            "evaluated_at": "2026-08-21T00:00:00+00:00",
                            "dataset_name": "unit-test",
                            "model": "gemini-test",
                            "provider": "vertex_ai",
                            "location": "global",
                            "prompt_type": prompt_type,
                            "prompt_version": version,
                            "metric_group": "common",
                            "metric": "success_rate",
                            "value": value,
                            "unit": "ratio",
                            "target": "",
                            "direction": "higher",
                            "passed": "",
                        }
                    )
            _write_csv(summary_path, summary_fields, summary_rows)

            comparison_fields = [
                "dataset_name",
                "model",
                "prompt_type",
                "metric_group",
                "metric",
                "v1_value",
                "v2_value",
                "unit",
                "target",
                "direction",
                "delta",
                "change_percent",
                "winner",
            ]
            comparison_rows = [
                {
                    "dataset_name": "unit-test",
                    "model": "gemini-test",
                    "prompt_type": prompt_type,
                    "metric_group": "common",
                    "metric": "success_rate",
                    "v1_value": 0.8,
                    "v2_value": 0.9,
                    "unit": "ratio",
                    "target": "",
                    "direction": "higher",
                    "delta": 0.1,
                    "change_percent": 12.5,
                    "winner": "v2",
                }
                for prompt_type in ("detection", "attribute")
            ]
            _write_csv(comparison_path, comparison_fields, comparison_rows)

            xai_fields = [
                "evaluated_at",
                "case_id",
                "prompt_version",
                "model",
                "success",
                "duration_ms",
                *prompt_excel_report.XAI_QUALITY_METRICS,
                "total_score",
                "match_status",
                "error",
            ]
            xai_rows = []
            for version, duration in (("v1", 1000), ("v2", 800)):
                xai_rows.append(
                    {
                        "evaluated_at": "2026-08-21T00:00:00+00:00",
                        "case_id": "chair-1",
                        "prompt_version": version,
                        "model": "gemini-test",
                        "success": 1,
                        "duration_ms": duration,
                        **{
                            metric: 1.0
                            for metric in prompt_excel_report.XAI_QUALITY_METRICS
                        },
                        "total_score": 80,
                        "match_status": "Matched",
                        "error": "",
                    }
                )
            _write_csv(xai_path, xai_fields, xai_rows)

            result = prompt_excel_report.generate_workbook(
                summary_path,
                comparison_path,
                xai_path,
                output_path,
            )

            self.assertEqual(result, output_path)
            workbook = load_workbook(output_path, data_only=False)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "종합 비교",
                    "공통 지표",
                    "객체 탐지",
                    "속성 추출",
                    "XAI",
                    "원본 결과",
                ],
            )
            self.assertTrue(
                str(workbook["종합 비교"]["A5"].value).startswith("=")
            )
            xai_values = [
                cell.value
                for row in workbook["XAI"].iter_rows()
                for cell in row
            ]
            self.assertIn("평균 처리 시간", xai_values)
            self.assertIn("v2", xai_values)


if __name__ == "__main__":
    unittest.main()
