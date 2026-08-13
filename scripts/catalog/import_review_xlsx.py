"""사람 검수용 XLSX를 최종 sku.json으로 변환한다.

실행:

python -m scripts.catalog.import_review_xlsx

입력:
data/catalog/answer/sku.json
data/catalog/review/sku_review.xlsx

출력:
data/catalog/answer/sku.json

검수 반영 규칙:

1. reviewed_value가 있으면 reviewed_value를 사용한다.
2. reviewed_value가 비어 있으면 기존 value를 사용한다.
3. 둘 다 비어 있으면 None을 사용한다.
4. selling_price는 숫자로 변환한다.
5. review_status / review_note / product_url은 최종 sku.json에 저장하지 않는다.
"""

from __future__ import annotations

import math
import pathlib
from typing import Any

from openpyxl import load_workbook

from scripts.catalog import storage


# ============================================================
# 경로
# ============================================================

REVIEW_XLSX_PATH = (
    storage.PROJECT_ROOT
    / "data"
    / "catalog"
    / "review"
    / "sku_review.xlsx"
)

OUTPUT_PATH = (
    storage.PROJECT_ROOT
    / "data"
    / "catalog"
    / "answer"
    / "sku.json"
)


# ============================================================
# XLSX 값 정리
# ============================================================

def normalize_cell_value(
    value: Any,
) -> Any:
    """Excel 셀 값을 JSON 저장에 적합한 값으로 정리한다."""

    # None
    if value is None:
        return None

    # 문자열
    if isinstance(value, str):
        value = value.strip()

        if value == "":
            return None

        return value

    # Excel에서 숫자가 float로 들어오는 경우
    if isinstance(value, float):
        if math.isnan(value):
            return None

        if value.is_integer():
            return int(value)

    return value


def convert_attribute_value(
    attribute: str,
    value: Any,
) -> Any:
    """속성별 최종 데이터 타입을 정리한다."""

    value = normalize_cell_value(value)

    if value is None:
        return None

    # 판매가는 숫자로 저장
    if attribute == "selling_price":
        if isinstance(value, str):
            # "66,900" 같은 형태 방어
            value = value.replace(",", "").strip()

            try:
                return int(float(value))
            except ValueError:
                return value

        if isinstance(value, float):
            return int(value)

        return value

    return value


# ============================================================
# 검수 XLSX 읽기
# ============================================================

def load_review_rows(
    path: pathlib.Path,
) -> list[dict[str, Any]]:
    """검수 XLSX의 SKU Review 시트를 읽는다."""

    if not path.exists():
        raise FileNotFoundError(
            f"검수 파일이 없습니다: {path}"
        )

    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
    )

    try:
        if "SKU Review" not in workbook.sheetnames:
            raise ValueError(
                "'SKU Review' 시트를 찾을 수 없습니다."
            )

        sheet = workbook["SKU Review"]

        rows = sheet.iter_rows(
            values_only=True
        )

        try:
            headers = next(rows)
        except StopIteration:
            return []

        headers = [
            str(header).strip()
            if header is not None
            else ""
            for header in headers
        ]

        result: list[dict[str, Any]] = []

        for row in rows:
            data = dict(
                zip(headers, row)
            )

            # 완전히 빈 행은 무시
            if not any(
                value is not None
                for value in data.values()
            ):
                continue

            result.append(data)

        return result

    finally:
        workbook.close()


# ============================================================
# SKU별 기존 데이터 인덱스
# ============================================================

def load_existing_skus() -> dict[int, dict[str, Any]]:
    """기존 sku.json을 sku_id 기준으로 읽는다."""

    existing = storage.load_json(
        storage.OUTPUT_PATH,
        [],
    )

    result: dict[int, dict[str, Any]] = {}

    for sku in existing:
        sku_id = sku.get("sku_id")

        if sku_id is None:
            continue

        result[int(sku_id)] = sku

    return result


# ============================================================
# 검수값 선택
# ============================================================

def select_final_value(
    value: Any,
    reviewed_value: Any,
) -> Any:
    """기존값과 검수값 중 최종값을 선택한다."""

    reviewed_value = normalize_cell_value(
        reviewed_value
    )

    value = normalize_cell_value(
        value
    )

    # 사람이 입력한 검수값이 있으면 최우선
    if reviewed_value is not None:
        return reviewed_value

    # 검수값이 없으면 기존값 유지
    return value


# ============================================================
# SKU 생성
# ============================================================

def build_final_skus(
    review_rows: list[dict[str, Any]],
    existing_skus: dict[int, dict[str, Any]],
) -> list[dict[str, Any]]:
    """검수 row를 최종 sku.json 구조로 합친다."""

    # SKU별 attributes를 임시 저장
    attributes_by_sku: dict[
        int,
        dict[str, Any],
    ] = {}

    # 기존 SKU 정보 복사
    final_skus: dict[
        int,
        dict[str, Any],
    ] = {}

    for sku_id, sku in existing_skus.items():

        # 기존 객체를 얕게 복사
        final_skus[sku_id] = {
            "sku_id": sku.get("sku_id"),
            "goods_id": sku.get("goods_id"),
            "sku_code": sku.get("sku_code"),
            "product_name": sku.get(
                "product_name",
                "",
            ),
            "category": sku.get(
                "category",
                "",
            ),
            "sub_category": sku.get(
                "sub_category",
                "",
            ),
            "key_features": sku.get(
                "key_features",
                [],
            ),
            "attributes": {},
        }

        attributes_by_sku[sku_id] = {}

    # ========================================================
    # 검수 row → attributes
    # ========================================================

    for row in review_rows:

        sku_id_raw = row.get("sku_id")

        if sku_id_raw is None:
            continue

        try:
            sku_id = int(sku_id_raw)
        except (
            TypeError,
            ValueError,
        ):
            print(
                f"[경고] 잘못된 sku_id: "
                f"{sku_id_raw}"
            )
            continue

        # 기존 sku.json에 없는 SKU
        if sku_id not in final_skus:
            print(
                f"[경고] 기존 sku.json에 "
                f"없는 sku_id: {sku_id}"
            )

            # 신규 SKU를 허용하려면 기본 정보 생성
            final_skus[sku_id] = {
                "sku_id": sku_id,
                "goods_id": row.get(
                    "goods_id"
                ),
                "sku_code": row.get(
                    "sku_code",
                    "",
                ),
                "product_name": row.get(
                    "product_name",
                    "",
                ),
                "category": row.get(
                    "category",
                    "",
                ),
                "sub_category": row.get(
                    "sub_category",
                    "",
                ),
                "key_features": [],
                "attributes": {},
            }

            attributes_by_sku[sku_id] = {}

        attribute = row.get(
            "attribute"
        )

        if not attribute:
            continue

        attribute = str(
            attribute
        ).strip()

        # 기존 value
        value = row.get(
            "value"
        )

        # 사람이 검수한 value
        reviewed_value = row.get(
            "reviewed_value"
        )

        final_value = select_final_value(
            value,
            reviewed_value,
        )

        final_value = convert_attribute_value(
            attribute,
            final_value,
        )

        attributes_by_sku[
            sku_id
        ][attribute] = final_value

    # ========================================================
    # attributes 적용
    # ========================================================

    for sku_id, sku in final_skus.items():

        sku["attributes"] = (
            attributes_by_sku.get(
                sku_id,
                {},
            )
        )

    # sku_id 순으로 정렬
    return [
        final_skus[sku_id]
        for sku_id in sorted(
            final_skus
        )
    ]


# ============================================================
# 저장
# ============================================================

def save_final_skus(
    skus: list[dict[str, Any]],
    path: pathlib.Path,
) -> None:
    """최종 sku.json을 저장한다."""

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    storage.dump_json(
        path,
        skus,
    )


# ============================================================
# Main
# ============================================================

def main() -> None:
    """검수 XLSX → 최종 sku.json."""

    print(
        "검수 XLSX 읽는 중..."
    )

    review_rows = load_review_rows(
        REVIEW_XLSX_PATH
    )

    print(
        f"검수 row: "
        f"{len(review_rows):,}건"
    )

    print(
        "기존 sku.json 읽는 중..."
    )

    existing_skus = load_existing_skus()

    print(
        f"기존 SKU: "
        f"{len(existing_skus):,}개"
    )

    final_skus = build_final_skus(
        review_rows,
        existing_skus,
    )

    save_final_skus(
        final_skus,
        OUTPUT_PATH,
    )

    relative_path = (
        OUTPUT_PATH.relative_to(
            storage.PROJECT_ROOT
        )
    )

    print()
    print(
        "========================================"
    )
    print(
        "최종 sku.json 생성 완료"
    )
    print(
        "========================================"
    )
    print(
        f"SKU 수: "
        f"{len(final_skus):,}개"
    )
    print(
        f"출력: "
        f"{relative_path}"
    )


if __name__ == "__main__":
    main()