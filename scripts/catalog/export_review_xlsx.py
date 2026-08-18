"""sku.json을 사람 검수용 Excel 파일로 변환한다.

실행:
    python -m scripts.catalog.export_review_xlsx

생성:
    data/catalog/review/sku_review.xlsx

Excel 구성:
    - SKU Review
        사람 검수용 메인 시트
    - Allowed Values
        속성별 허용값 확인용 시트
    - Dropdown Values
        reviewed_value 드롭다운용 숨김 시트

검수 방식:
    - 허용값이 정의된 속성:
        reviewed_value -> 드롭다운 선택
    - 허용값이 없는 속성:
        reviewed_value -> 직접 입력
    - review_status:
        pending / approved / corrected
    - review_note:
        검수자가 자유롭게 메모
"""

from __future__ import annotations

import pathlib
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

from app.core import catalog_spec
from scripts.catalog import storage


# ============================================================
# 경로
# ============================================================

OUTPUT_PATH = (
    storage.PROJECT_ROOT
    / "data"
    / "catalog"
    / "review"
    / "sku_review.xlsx"
)


# ============================================================
# Excel 컬럼
# ============================================================

REVIEW_HEADERS = [
    "sku_id",
    "goods_id",
    "sku_code",
    "product_url",
    "product_name",
    "category",
    "sub_category",
    "attribute",
    "value",
    "reviewed_value",
    "review_status",
    "review_note",
]


REVIEW_STATUSES = [
    "pending",
    "approved",
    "corrected",
]


# ============================================================
# SKU -> 검수 row
# ============================================================

def build_review_rows(
    skus: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """SKU 데이터를 속성 단위 검수 row로 펼친다."""

    rows: list[dict[str, Any]] = []

    for sku in skus:
        attributes = sku.get("attributes") or {}

        goods_id = sku.get("goods_id")

        # 오늘의집 상품 URL
        product_url = (
            f"https://ohou.se/productions/{goods_id}/selling"
            if goods_id is not None
            else ""
        )

        for attribute, value in attributes.items():
            rows.append(
                {
                    "sku_id": sku.get("sku_id"),
                    "goods_id": goods_id,
                    "sku_code": sku.get(
                        "sku_code",
                        "",
                    ),
                    "product_url": product_url,
                    "product_name": sku.get(
                        "product_name",
                        "",
                    ),
                    "category": sku.get(
                        "category",
                        "",
                    ),
                    "sub_category": sku.get(
                        "sub_category",
                        "",
                    ),
                    "attribute": attribute,
                    "value": value,
                    "reviewed_value": "",
                    "review_status": "pending",
                    "review_note": "",
                }
            )

    return rows


# ============================================================
# 허용값 수집
# ============================================================

def build_allowed_values(
    skus: list[dict[str, Any]],
) -> dict[tuple[str, str], list[str]]:
    """카테고리 + 속성별 허용값을 만든다.

    반환 예:

        {
            ("서랍·수납장", "color"):
                ["블랙", "화이트", ...],

            ("서랍·수납장", "drawer_count"):
                ["1단", "2단", ...],
        }

    brand / selling_price처럼 허용값이 없는 속성은
    결과에 포함하지 않는다.
    """

    result: dict[
        tuple[str, str],
        set[str],
    ] = {}

    for sku in skus:
        category = sku.get("category")

        if not category:
            continue

        attributes = sku.get("attributes") or {}

        for attribute in attributes:
            try:
                allowed = catalog_spec.allowed_values(
                    category,
                    attribute,
                )
            except Exception:
                continue

            if not allowed:
                continue

            key = (
                category,
                attribute,
            )

            result.setdefault(
                key,
                set(),
            ).update(
                str(value)
                for value in allowed
                if value is not None
            )

    return {
        key: sorted(values)
        for key, values in result.items()
    }


# ============================================================
# XLSX 생성
# ============================================================

def write_xlsx(
    rows: list[dict[str, Any]],
    allowed_values: dict[
        tuple[str, str],
        list[str],
    ],
    path: pathlib.Path,
) -> None:
    """검수용 XLSX 파일을 생성한다."""

    path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    # ========================================================
    # 1. SKU Review
    # ========================================================

    review_sheet = workbook.active
    review_sheet.title = "SKU Review"

    # 헤더
    review_sheet.append(
        REVIEW_HEADERS
    )

    # 데이터
    for row in rows:
        review_sheet.append(
            [
                row.get(header)
                for header in REVIEW_HEADERS
            ]
        )

    # 첫 행 고정
    review_sheet.freeze_panes = "A2"

    # 필터
    last_column = get_column_letter(
        len(REVIEW_HEADERS)
    )

    review_sheet.auto_filter.ref = (
        f"A1:{last_column}{len(rows) + 1}"
    )

    # ========================================================
    # URL 하이퍼링크
    # ========================================================

    # product_url = D열
    for row_number in range(
        2,
        len(rows) + 2,
    ):
        cell = review_sheet.cell(
            row=row_number,
            column=4,
        )

        if cell.value:
            cell.hyperlink = str(
                cell.value
            )

            cell.style = "Hyperlink"

    # ========================================================
    # 2. Allowed Values
    # ========================================================

    allowed_sheet = workbook.create_sheet(
        "Allowed Values"
    )

    allowed_sheet.append(
        [
            "category",
            "attribute",
            "allowed_value",
        ]
    )

    for (
        category,
        attribute,
    ), values in sorted(
        allowed_values.items()
    ):
        for value in values:
            allowed_sheet.append(
                [
                    category,
                    attribute,
                    value,
                ]
            )

    allowed_sheet.freeze_panes = "A2"

    # ========================================================
    # 3. Dropdown Values
    # ========================================================

    # Excel의 데이터 유효성 검사에서 사용할
    # 실제 드롭다운 목록을 저장하는 시트
    dropdown_sheet = workbook.create_sheet(
        "Dropdown Values"
    )

    # --------------------------------------------------------
    # 각 (category, attribute)별로 하나의 열을 만든다.
    #
    # 예:
    #
    # A열
    # 서랍·수납장__color
    # 블랙
    # 화이트
    # ...
    #
    # B열
    # 서랍·수납장__drawer_count
    # 1단
    # 2단
    # ...
    # --------------------------------------------------------

    dropdown_keys = sorted(
        allowed_values.keys()
    )

    defined_names: dict[
        tuple[str, str],
        str,
    ] = {}

    for column_index, key in enumerate(
        dropdown_keys,
        start=1,
    ):
        category, attribute = key

        column_letter = get_column_letter(
            column_index
        )

        # Excel 헤더
        header = (
            f"{category}__{attribute}"
        )

        dropdown_sheet.cell(
            row=1,
            column=column_index,
            value=header,
        )

        values = allowed_values[key]

        for row_index, value in enumerate(
            values,
            start=2,
        ):
            dropdown_sheet.cell(
                row=row_index,
                column=column_index,
                value=value,
            )

        # ----------------------------------------------------
        # Excel Named Range
        # ----------------------------------------------------

        last_row = len(values) + 1

        # 이름은 category/attribute의 한글,
        # 특수문자 등에 영향을 받지 않도록
        # 단순한 번호 기반으로 만든다.
        defined_name = (
            f"dv_{column_index}"
        )

        defined_names[key] = defined_name

        name = DefinedName(
            defined_name,
            attr_text=(
                f"'Dropdown Values'!"
                f"${column_letter}$2:"
                f"${column_letter}${last_row}"
            ),
        )

        workbook.defined_names.add(
            name
        )

    # ========================================================
    # 4. reviewed_value 드롭다운
    # ========================================================

    # SKU Review의 컬럼 위치
    #
    # A sku_id
    # B goods_id
    # C sku_code
    # D product_url
    # E product_name
    # F category
    # G sub_category
    # H attribute
    # I value
    # J reviewed_value
    # K review_status
    # L review_note

    for row_number, row in enumerate(
        rows,
        start=2,
    ):
        category = row.get(
            "category"
        )

        attribute = row.get(
            "attribute"
        )

        if not category or not attribute:
            continue

        key = (
            category,
            attribute,
        )

        # 허용값이 없는 속성
        #
        # 예:
        # brand
        # selling_price
        #
        # -> 직접 입력
        if key not in defined_names:
            continue

        defined_name = defined_names[key]

        validation = DataValidation(
            type="list",
            formula1=f"={defined_name}",
            allow_blank=True,
        )

        validation.error = (
            "허용된 값 중 하나를 선택하세요."
        )

        validation.errorTitle = (
            "허용되지 않은 값"
        )

        validation.prompt = (
            f"{attribute}의 허용값을 "
            "선택하세요."
        )

        validation.promptTitle = (
            "검수값 선택"
        )

        review_sheet.add_data_validation(
            validation
        )

        # J열 = reviewed_value
        validation.add(
            f"J{row_number}"
        )

    # ========================================================
    # 5. review_status 드롭다운
    # ========================================================

    status_validation = DataValidation(
        type="list",
        formula1='"pending,approved,corrected"',
        allow_blank=False,
    )

    status_validation.error = (
        "pending, approved, corrected 중 "
        "하나를 선택하세요."
    )

    status_validation.errorTitle = (
        "잘못된 검수 상태"
    )

    status_validation.prompt = (
        "검수 상태를 선택하세요."
    )

    status_validation.promptTitle = (
        "검수 상태"
    )

    review_sheet.add_data_validation(
        status_validation
    )

    if rows:
        # K열 = review_status
        status_validation.add(
            f"K2:K{len(rows) + 1}"
        )

    # ========================================================
    # 6. 컬럼 너비
    # ========================================================

    widths = {
        "A": 10,   # sku_id
        "B": 12,   # goods_id
        "C": 20,   # sku_code
        "D": 50,   # product_url
        "E": 50,   # product_name
        "F": 20,   # category
        "G": 20,   # sub_category
        "H": 22,   # attribute
        "I": 22,   # value
        "J": 22,   # reviewed_value
        "K": 18,   # review_status
        "L": 40,   # review_note
    }

    for column, width in widths.items():
        review_sheet.column_dimensions[
            column
        ].width = width

    # Allowed Values
    allowed_sheet.column_dimensions[
        "A"
    ].width = 20

    allowed_sheet.column_dimensions[
        "B"
    ].width = 25

    allowed_sheet.column_dimensions[
        "C"
    ].width = 30

    # Dropdown Values
    for column_index in range(
        1,
        len(dropdown_keys) + 1,
    ):
        dropdown_sheet.column_dimensions[
            get_column_letter(column_index)
        ].width = 25

    # 검수자가 직접 수정할 필요 없는 시트
    dropdown_sheet.sheet_state = "hidden"

    # ========================================================
    # 저장
    # ========================================================

    workbook.save(path)


# ============================================================
# Main
# ============================================================

def main() -> None:
    """sku.json을 검수용 XLSX로 변환한다."""

    skus = storage.load_json(
        storage.OUTPUT_PATH,
        [],
    )

    if not skus:
        print(
            "sku.json에 SKU 데이터가 없습니다."
        )
        return

    rows = build_review_rows(
        skus
    )

    allowed_values = build_allowed_values(
        skus
    )

    write_xlsx(
        rows,
        allowed_values,
        OUTPUT_PATH,
    )

    relative_path = (
        OUTPUT_PATH.relative_to(
            storage.PROJECT_ROOT
        )
    )

    print(
        f"검수 XLSX 생성 완료: "
        f"{relative_path}"
    )

    print(
        f"검수 대상 행: "
        f"{len(rows)}건"
    )

    print(
        f"드롭다운 속성: "
        f"{len(allowed_values)}개"
    )


# ============================================================
# Entry point
# ============================================================

if __name__ == "__main__":
    main()